'''
Laddar ner senaste datan från Socialstyrelsen på avlidna och sjukhusinläggningar i Covid-19.
'''
from urllib.request import urlretrieve
from openpyxl import load_workbook
import pandas as pd

# Uppdaterar till senaste versionen
url = 'https://www.socialstyrelsen.se/globalassets/1-globalt/covid-19-statistik/vard-och-covid-19/statistik-covid19-inskrivna.xlsx'
urlretrieve(url, 'Covid19_sjukhus.xlsx')

# Öppnar arbetsboken
wb = load_workbook('Covid19_sjukhus.xlsx')
ws = wb['Inskrivna i slutenvård']

# Plockar ut datan vi behöver från filerna
data = []
row = [cell.value for cell in ws[8]]
for i in range(3, len(row)):
    if i % 2 != 0:
        try:
            data.append(row[i]+data[-1])
        except IndexError:
            data.append(row[i])

week1 = [x for x in range(10,54)]
week1 = week1 + [x for x in range(1,len(data)-len(week1)+1)]

# Lägger in datan i en ny csv-fil
dict = {'vecka': week1, 'kum_antal_sjukhus': data}
df = pd.DataFrame(dict)
df.to_csv('Covid19_sjukhus.csv')


## Samma sak fast för antal avlidna: ###
# Uppdaterar till senaste versionen
url2 = 'https://www.socialstyrelsen.se/globalassets/1-globalt/covid-19-statistik/avlidna-i-covid-19/statistik-covid19-avlidna.xlsx'
urlretrieve(url2, 'Covid19_avlidna.xlsx')

# Öppnar arbetsboken
wb2 = load_workbook('Covid19_avlidna.xlsx')
ws2 = wb2['Vecka']

# Plockar ut datan vi behöver
data2 = []
row2 = [cell.value for cell in ws2['B']]
for i in range(10, len(row2)):
    if i < 12 and type(row2[i]) != int:
        data2.append(0)
    elif i >= 12 and type(row2[i]) != int:
        break
    else:
        data2.append(row2[i]+data2[-1])
        
week = [x for x in range(10,54)]
week = week + [x for x in range(1,len(data2)-len(week)+1)]

# Lägger in datan i en ny csv-fil
dict = {'vecka': week, 'kum_antal_avlidna': data2}
df = pd.DataFrame(dict)
df.to_csv('Covid19_avlidna.csv')